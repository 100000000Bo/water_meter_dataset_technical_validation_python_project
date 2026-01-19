import os
import shutil
import pandas as pd
from openpyxl import load_workbook


def rename_images_and_excel(excel_path, image_folder):
    # Load the Excel file using openpyxl to preserve formatting
    wb = load_workbook(excel_path)
    ws = wb.active

    # Ensure that the first column exists
    if ws.max_column < 1:
        print("The Excel file must contain at least one column.")
        return

    # Read image file names from the first column (skip header row)
    image_names = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]

    # Generate new image file names with consistent indexing
    new_names = [f"train{i}" + os.path.splitext(str(name))[1] for i, name in enumerate(image_names)]

    for old_name, new_name in zip(image_names, new_names):
        old_path = os.path.join(image_folder, str(old_name))
        new_path = os.path.join(image_folder, new_name)

        if os.path.exists(old_path):
            shutil.move(old_path, new_path)
        else:
            print(f"File not found: {old_path}")

    # Update the first column in the Excel file with new image names
    for i, new_name in enumerate(new_names, start=2):
        ws.cell(row=i, column=1, value=new_name)

    # Save the updated Excel file while preserving formatting
    wb.save(excel_path)
    print("Renaming completed. The Excel file has been updated.")


rename_images_and_excel(
    "C:/Users/yibo/Desktop/train1.xlsx",
    "C:/Users/yibo/Desktop/6-train"
)
